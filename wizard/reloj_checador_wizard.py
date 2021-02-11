# -*- encoding: utf-8 -*-

from odoo import models, fields, api, _
from odoo.exceptions import ValidationError, UserError, Warning
import time
import base64
import xlwt
import io
import logging
import requests
import json
from datetime import date
import datetime
import datetime
from datetime import datetime
import glob
import time
import openpyxl

import base64
import xlsxwriter
import io
import logging
import xlrd
import base64

class QuemenRelojChecadorWizard(models.TransientModel):
    _name = 'quemen.reloj_checador.wizard'

    archivo = fields.Binary('Archivo excel')

    def _get_horas(self, horas):
        hora_entrada = 0
        hora_salida = 0
        dic = {'hora_entrada': 0,'hora_salida':0}
        if horas:
            horas = horas.split()
            tamanio_horas = len(horas)
            if tamanio_horas == 1:
                hora_entrada =  horas[0]

            elif tamanio_horas > 1:
                hora_entrada =  horas[0]
                hora_salida = horas[tamanio_horas-1]
            # logging.warn(horas[])
            # logging.warn(horas.split())
            dic['hora_entrada'] = hora_entrada
            dic['hora_salida'] = hora_salida
        return dic

    def cargar_datos(self):
        workbook = xlrd.open_workbook(file_contents = base64.decodestring(self.archivo))
        sheet = workbook.sheet_by_index(0)
        empleados = self.env['hr.employee'].search([])
        empleados_dic = {}
        for empleado in empleados:
            empleados_dic[empleado.id_reloj] = {'empleado': empleado}

        logging.warn('emp')
        logging.warn(empleados_dic)
        for linea in range(sheet.nrows):
            if linea != 0:
                ac = sheet.cell(linea, 0).value
                fecha_excel = sheet.cell(linea, 3).value
                horas = sheet.cell(linea, 4).value
                hora_entrada = self._get_horas(horas)['hora_entrada']
                hora_salida = self._get_horas(horas)['hora_salida']
                if str(ac) in empleados_dic:
                    empleado = empleados_dic[str(ac)]

                    if hora_entrada != 0:
                        hora_entrada_split = hora_entrada.split(":")
                        hora = float(hora_entrada_split[0])
                        minuto = float(hora_entrada_split[1])
                        hora_entrada = float(hora + minuto / 60)
                    if hora_salida != 0:
                        hora_salida_split = hora_salida.split(":")
                        hora = float(hora_salida_split[0])
                        minuto = float(hora_salida_split[1])
                        hora_salida = float(hora + minuto / 60)

                    reloj =                     {'empleado_id': empleado['empleado'].id,
                                        'hora_entrada': hora_entrada,
                                        'hora_salida': hora_salida,
                                        'fecha': datetime.strptime(str(fecha_excel), "%d/%m/%Y").date(),
                                        }
                    logging.warn(reloj)
                    reloj_id = self.env['quemen.reloj_checador'].create(reloj)
                else:
                    raise ValidationError(_("codigo no existe"+" "+str(ac)))

        return {
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'quemen.reloj_checador.wizard',
            'res_id': self.id,
            'view_id': False,
            'type': 'ir.actions.act_window',
            'target': 'new',
        }
